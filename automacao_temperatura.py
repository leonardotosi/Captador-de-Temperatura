# Importa as Bibliotecas necessárias
import time
import os
from datetime import datetime
from tkinter import Tk, Label, Button, StringVar, ttk, messagebox
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook

# Retorna o WebDriver do Selenium e Inicializa o Google
def inicializar_navegador():
   
    return webdriver.Chrome()

#Acessa o Google
def acessar_google(navegador):
    
    navegador.get('https://www.google.com')

# Busca a previsão
def buscar_previsao_tempo(navegador):
    
    navegador.find_element(By.XPATH, '//*[@id="APjFqb"]').send_keys('previsao tempo')
    navegador.find_element(By.XPATH, '//*[@id="APjFqb"]').send_keys(Keys.ENTER)
    time.sleep(2)  # Aguarda o carregamento da página

# Obtém os dados de data, temperatura, umidade e condição via ID
def obter_dados_previsao(navegador):
    
    try:
        temperatura = navegador.find_element(By.ID, "wob_tm").text
        umidade = navegador.find_element(By.ID, "wob_hm").text
        condicao = navegador.find_element(By.ID, "wob_dc").text
    except Exception as e:
        print(f"Erro ao obter os dados: {e}")
        return None, None, None

    return temperatura, umidade, condicao

# Salva no Excel
def salvar_dados_no_arquivo(temperatura, umidade, condicao, nome_arquivo='historico_temperatura.xlsx'):
    
    try:
        # Verifica se o arquivo já existe
        if os.path.exists(nome_arquivo):
            arquivo = load_workbook(nome_arquivo)
        else:
            arquivo = Workbook()

        planilha = arquivo.active

        if planilha.max_row == 1:
            planilha.append(['Data', 'Temperatura', 'Umidade', 'Condição'])

        planilha.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), temperatura, umidade, condicao])

        for col in planilha.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            planilha.column_dimensions[col[0].column_letter].width = max_length + 2

        arquivo.save(nome_arquivo)
        return f"Dados salvos em {nome_arquivo}."
    except Exception as e:
        return f"Erro ao salvar os dados: {e}"

# Função principal que executa toda a tarefa
def executar_tarefa():
    
    try:
        navegador = inicializar_navegador()
        acessar_google(navegador)
        buscar_previsao_tempo(navegador)
        temperatura, umidade, condicao = obter_dados_previsao(navegador)

        if temperatura and umidade and condicao:
            resultado = salvar_dados_no_arquivo(temperatura, umidade, condicao)
            messagebox.showinfo("Sucesso", resultado)
        else:
            messagebox.showwarning("Atenção", "Falha ao obter os dados.")

        navegador.quit()
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Interface Gráfica com tkinter
root = Tk()
root.title("Previsão do Tempo")
root.geometry("400x300")
root.resizable(False, False)

# Tema do tkinter
style = ttk.Style()
style.theme_use("clam")
style.configure("TButton", font=("Segoe UI", 12), padding=6)
style.configure("TLabel", font=("Segoe UI", 11))


frame_titulo = ttk.Frame(root, padding=10)
frame_titulo.pack(fill="x")

titulo = ttk.Label(frame_titulo, text="Captura de Previsão do Tempo", font=("Segoe UI", 16, "bold"), anchor="center")
titulo.pack()

frame_conteudo = ttk.Frame(root, padding=20)
frame_conteudo.pack(fill="both", expand=True)

botao_iniciar = ttk.Button(frame_conteudo, text="Obter Dados", command=executar_tarefa)
botao_iniciar.pack(pady=20)

status_label = ttk.Label(frame_conteudo, text="Clique no botão para iniciar.", anchor="center")
status_label.pack(pady=10)

# Roda a aplicação
root.mainloop()